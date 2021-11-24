import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook


#Leemos en TXT
print('Iniciando el proceso...')
print('Leyendo archivo fuente TXT...')
datos = pd.read_csv('input/Colores.txt', sep=None, engine="python")

#Guardamos en Excel en DataFrame cargado del archivo TXT
print('Guardando en archivo de salida Excel hoja Colores...')
datos.to_excel('output/reporte.xlsx',sheet_name='Colores', index=False,
                engine = 'openpyxl')

#Leemos en CSV
print('Leyendo archivo fuente CSV...')
datos1 = pd.read_csv('input/Ciudades.csv', sep=";")

#Guardamos en Excel en DataFrame cargado del archivo CSV
print('Guardando en archivo de salida Excel hoja Ciudades...')
with pd.ExcelWriter('output/reporte.xlsx', mode='a') as writer:
                    datos1.to_excel(writer, sheet_name='Ciudades', index=False)
print('El proceso ha finalizado con exito.')


#Creamos Libro
#libro = Workbook()
#libro.save('output/reporte.xlsx')

#Leemos archivo
#clibro = load_workbook('output/reporte.xlsx')

#Creamos Hojas
#hoja1 = clibro.create_sheet("Casas",0)
#hoja1.title = "CASAS"
#hoja2 = clibro.create_sheet("Colores",1)

#Guardamos los cambios
#clibro.save('output/reporte.xlsx')

#df = pd.read_excel(r'C:\Users\lufeg\Documents\python\pandas\input\Plantilla.xlsx')

#print(df.head(3))

#wb2 = load_workbook(r'C:\Users\lufeg\Documents\python\pandas\input\Plantilla.xlsx')

#ws1 = wb2.create_sheet('Hoja1001')

#wb2.save(r'C:\Users\lufeg\Documents\python\pandas\input\Plantilla.xlsx')
#df.to_excel("C:\\Users\\lufeg\\Documents\\python\\pandas\\output\\PlantillaR.xlsx")